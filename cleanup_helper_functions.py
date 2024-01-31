import pandas as pd
import numpy as np
import openpyxl
import os
import ast


def normalize_string(string_in):
    # Normalize Names and Addresses: lowercase, strip whitespace, replace multiple whitespace with single whitespace
    normalized = string_in.lower().strip()
    normalized = re.sub(r"\s+", " ", normalized)
    return normalized


def basic_cleanup(df, organisation=False):
    """
    Performs some basic corrections to String formatting.
    Removes Inaktiv entries and Personen with Sonstiges Verknüpfungsart.
    """

    # Remove rows missing a name.
    df_cleaned = df[df["Name"].apply(lambda x: isinstance(x, str))]

    df_cleaned["Name_original"] = df_cleaned["Name"]  # Keep original for reference
    df_cleaned["Name"] = df_cleaned["Name"].apply(normalize_string)

    # sort by Name
    df_cleaned = df_cleaned.sort_values("Name")

    df_cleaned = df_cleaned.replace(
        {pd.NA: "", "nan": ""}
    )  # sometimes cells contain string 'nan' this causes problems later

    # Remove this step, because we want to treat "NotRegisteredCHId" separately from nan.
    # df_cleaned.replace("NotRegisteredCHId", pd.NA, inplace=True)  # also for ch-uid
    # df_cleaned.replace("NotRegisteredCHID", pd.NA, inplace=True)  # also for ch-uid

    # BUG: Remember that .astype(str) will replace pd.NA with "nan".

    # removing all spaces between numbers, because they are placed quite inconsistently.
    df_cleaned["Telefonnummer"] = (
        df_cleaned["Telefonnummer"].str.replace(" ", "").astype(str)
    )
    # email addresses will also need some processing. for now, ensure values are strings
    df_cleaned["EMailAdresse"] = df_cleaned["EMailAdresse"].astype(str).str.lower()

    # Ensure that PLZ is string, for some reason its sometimes float which causes problems
    df_cleaned["ZipPostalCode"] = df_cleaned["ZipPostalCode"].astype(str)

    # Filter out all Inaktive and Sonstiges for Personen, to speed up processing
    df_cleaned = df_cleaned[df_cleaned["Aktiv"] != False]
    if not organisation:
        df_cleaned = df_cleaned[df_cleaned["Verknuepfungsart"] != "Sonstiges"]

    return df_cleaned


def construct_address_string(row, organisation=False):
    """
    expects row to have the elements listed below.
    Since extra text in address1 and address2 can confuse gmaps, also return partial address with only street and number.
    """
    # Check if ZipPostalCode is a number
    # Check if ZipPostalCode is not NaN and not the string 'nan'
    zip_code = row["ZipPostalCode"]
    if pd.notna(zip_code) and str(zip_code).lower() != "nan" and zip_code != "":
        try:
            zip_postal_code = str(int(float(zip_code)))
        except ValueError:
            zip_postal_code = str(zip_code)  # if it has letters, e.g. UK
    elif organisation:
        korr_zip_code = row["Korr_ZipPostalCode"]
        if (
            pd.notna(korr_zip_code)
            and str(korr_zip_code).lower() != "nan"
            and korr_zip_code != ""
        ):
            try:
                zip_postal_code = str(int(float(korr_zip_code)))
            except ValueError:
                zip_postal_code = str(korr_zip_code)
        else:
            zip_postal_code = ""
    else:
        zip_postal_code = ""

    address_elements = [
        str(row["Street"]),
        str(row["HouseNumber"]),
        str(row["Address1"]),
        str(row["Address2"]),
        str(row["PostOfficeBox"]),
        zip_postal_code,
        str(row["City"]),
        str(row["CountryName"]),
    ]

    address_elements_partial = [
        str(row["Street"]),
        str(row["HouseNumber"]),
        zip_postal_code,
        str(row["City"]),
        str(row["CountryName"]),
    ]

    elements_without_zip_code = [
        str(row["Street"]),
        str(row["HouseNumber"]),
        str(row["Address1"]),
        str(row["Address2"]),
        str(row["PostOfficeBox"]),
        str(row["City"]),
        str(row["CountryName"]),
    ]

    # Check if all address elements are NaN (or "nan" or empty strings), try use korrespondenz_adresse instead
    if all(
        pd.isna(element) or element == "" or element.lower() == "nan"
        for element in elements_without_zip_code
    ):
        if organisation:  # Personen don't have these columns
            address_elements = [
                str(row["Korr_Street"]),
                str(row["Korr_HouseNumber"]),
                str(row["Korr_Address1"]),
                str(row["Korr_Address2"]),
                str(row["Korr_PostOfficeBox"]),
                zip_postal_code,
                str(row["Korr_City"]),
                str(row["Korr_CountryName"]),
            ]

            address_elements_partial = [
                str(row["Korr_Street"]),
                str(row["Korr_HouseNumber"]),
                zip_postal_code,
                str(row["Korr_City"]),
                str(row["Korr_CountryName"]),
            ]
        else:
            return ""
        # it that is also empty, return empty string
        if all(
            pd.isna(element) or element == "" or element.lower() == "nan"
            for element in address_elements
        ):
            return ""

    # Filter out None, 'nan', and empty strings, then join with commas
    full_address = ", ".join(
        filter(lambda x: x and x != "nan" and str(x).strip(), address_elements)
    )

    partial_address = ", ".join(
        filter(lambda x: x and x != "nan" and str(x).strip(), address_elements_partial)
    )

    # Finally make it lowercase, remove additional spaces
    full_address = normalize_string(full_address)
    partial_address = normalize_string(partial_address)

    return pd.Series([full_address, partial_address])


def replace_NotRegisteredUID(df):
    # TODO: This should all be done in the cleaning notebook
    df.replace("NotRegisteredCHID", pd.NA, inplace=True)

    # Also fix some other weird ocurrences
    df = df.replace({np.nan: pd.NA})

    return df


def aggregate_identical_UIDs(df):
    """
    Those with identical IDs that are a result of flattening from linq output.
    Does not affect duplicated with same name and different IDs.
    VerknüpftesObjekt etc. are aggregated as lists.
    """

    def aggregate_to_list(series):
        return series.tolist()

    def first_entry(series):
        return series.iloc[0]

    columns_as_lists = [
        "Verknuepfungsart",
        "VerknuepftesObjektID",
        "VerknuepftesObjekt",
    ]

    aggregation = {
        col: aggregate_to_list if col in columns_as_lists else first_entry
        for col in df.columns
        if col != "ReferenceID"
    }

    grouped = df.groupby("ReferenceID").agg(aggregation).reset_index()

    return grouped


def extract_hyperlinks(file_path, columns):
    """
    Extract hyperlinks from specified columns in an Excel file and add new columns with the suffix `_link`.

    Parameters:
        file_path (str): Path to the Excel file.
        columns (list): List of column names to extract hyperlinks from.

    Returns:
        None. The function saves the updated DataFrame to a new file with a `_hyperlinks` suffix.
    """
    # Load the Excel file using openpyxl
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active

    # Dictionary to store hyperlinks for each column
    hyperlink_dict = {}
    for column_name in columns:
        col_idx = None
        for idx, col in enumerate(sheet.iter_cols(1, sheet.max_column)):
            if col[0].value == column_name:
                col_idx = idx + 1  # 1-based index
                break

        if col_idx is None:
            print(f"Column '{column_name}' not found.")
            continue

        # Extract hyperlinks
        hyperlinks = []
        for row in sheet.iter_rows(
            min_row=2, max_row=sheet.max_row, min_col=col_idx, max_col=col_idx
        ):
            cell = row[0]
            if cell.hyperlink:
                hyperlinks.append(cell.hyperlink.target)
            else:
                hyperlinks.append(None)

        hyperlink_dict[column_name + "_link"] = hyperlinks

    # Load the Excel data into pandas DataFrame, then add the hyperlinks
    df = pd.read_excel(file_path, engine="openpyxl")
    for col_name, links in hyperlink_dict.items():
        df[col_name] = links

    # Determine the save path
    base_name = os.path.basename(file_path)
    name_without_extension = os.path.splitext(base_name)[0]
    save_name = name_without_extension + "_hyperlinks.xlsx"
    save_path = os.path.join(os.path.dirname(file_path), save_name)

    # Save the updated DataFrame back to Excel with the new name
    df.to_excel(save_path, index=False, engine="openpyxl")

    return save_path


def extract_hyperlinks_optimized(file_path, columns):
    # performance optmized by phind. to be tested
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active

    hyperlink_dict = {}
    for column_name in columns:
        col_idx = None
        for idx, col in enumerate(sheet.iter_cols(1, sheet.max_column)):
            if col[0].value == column_name:
                col_idx = idx + 1
                break

        if col_idx is None:
            print(f"Column '{column_name}' not found.")
            continue

        hyperlinks = []
        for row in sheet.iter_rows(
            min_row=2,
            max_row=sheet.max_row,
            min_col=col_idx,
            max_col=col_idx,
            values_only=True,
        ):
            if (
                isinstance(row[0], str) and "http" in row[0]
            ):  # check if row[0] is a string and contains 'http'
                hyperlinks.append(row[0])
            else:
                hyperlinks.append(None)

        hyperlink_dict[column_name + "_link"] = hyperlinks

    df = pd.read_excel(file_path, engine="openpyxl")
    for col_name, links in hyperlink_dict.items():
        df[col_name] = links

    base_name = os.path.basename(file_path)
    name_without_extension = os.path.splitext(base_name)[0]
    save_name = name_without_extension + "_hyperlinks.xlsx"
    save_path = os.path.join(os.path.dirname(file_path), save_name)

    df.to_excel(save_path, index=False, engine="openpyxl")

    return save_path


def get_flat_list(key_list):
    # Solves a problem with the lists of "VerknuepftesObjektID" that are somehow strings looking like lists...
    flat_list = []
    for nested_string in key_list:
        inner_string = nested_string.strip("[]'")  # Remove brackets and single quotes
        flat_list.append(inner_string)
    # print(flat_list)
    return flat_list


def get_additional_organizations(
    df_filtered, df_original, check_columns="all", include_address=False
):
    """
    df_filtered is the list of organisations that is directly connected to a given person.
    df_original is the list of all organisations.
    If there is an organisation in df_original that matches an organisation in df_filtered,
    by comparing "columns_to_check", it is added to the results.
    A new column "match_type" is added showing what type of match was found.
    VerknuepftesObjektID is special, it looks if any of those IDs matches a ReferenceID of an organisation.
    Multiple entries are possible if two organisations match on more than one value.

    Note: Despite the function name is now also used to find additional Personen from Organisationen.
    """
    if check_columns == "all":
        columns_to_check = {
            "EMailAdresse": "Email",
            "Telefonnummer": "Telefon",
            "VerknuepftesObjektID": "Sonstiges",  # no longer hard-coded uses Verknüpfungsart value
        }
        if include_address:
            columns_to_check["address_gmaps"] = "Adresse"

    elif check_columns == "email_phone":
        columns_to_check = {
            "EMailAdresse": "Email",
            "Telefonnummer": "Telefon",
        }
        if include_address:
            columns_to_check[
                "address_gmaps"
            ] = "Adresse"  # not meaningful when indpenden companies are in same building

    elif check_columns == "ID_only":
        columns_to_check = {"VerknuepftesObjektID": "Sonstiges"}

    result = pd.DataFrame()

    for col, match_type in columns_to_check.items():
        if col != "VerknuepftesObjektID":
            # Filter rows based on the column length for 'address_gmaps'
            if col == "address_gmaps":
                temp_filtered = df_filtered[
                    (df_filtered[col].str.len() > 20) & df_filtered[col].notna()
                ][[col, "ReferenceID"]]
            else:
                temp_filtered = df_filtered[df_filtered[col].notna()][
                    [col, "ReferenceID"]
                ]
            temp = df_original.merge(
                temp_filtered, left_on=col, right_on=col, how="inner"
            )
            temp["match_type"] = match_type  # Add the match type
            temp.rename(
                columns={"ReferenceID_x": "ReferenceID", "ReferenceID_y": "source"},
                inplace=True,
            )
            result = pd.concat([result, temp], axis=0).drop_duplicates()
        else:
            # Loop through elements in the VerknuepftesObjektID column of df_filtered
            for _, row in df_filtered.iterrows():
                verknuepftes_objekt_ids = row["VerknuepftesObjektID"]

                try:
                    verknuepftes_objekt_ids = ast.literal_eval(verknuepftes_objekt_ids)
                except (SyntaxError, ValueError):
                    # Handle the error or skip the row as needed
                    continue

                verknuepfungsart_values = (
                    ast.literal_eval(row["Verknuepfungsart"])
                    if "Verknuepfungsart" in df_filtered.columns
                    and pd.notna(row["Verknuepfungsart"])
                    else [match_type] * len(verknuepftes_objekt_ids)
                )

                for item, match_val in zip(
                    verknuepftes_objekt_ids, verknuepfungsart_values
                ):
                    mask = (
                        df_original["ReferenceID"] == item
                    )  # Check if ReferenceID matches the item
                    temp = df_original[mask].copy()
                    temp[
                        "match_type"
                    ] = match_val  # Set the corresponding "Verknuepfungsart" value
                    temp["source"] = row["ReferenceID"]
                    result = pd.concat([result, temp], axis=0).drop_duplicates()

    return result


def find_internal_matches(df):
    from graphviz_helper_functions import convert_string_to_list

    """
    Very similar to function above (get_additional_organizations), but for a single dataframe.
    Warning: code duplication. Beware in case of changes.
    Intended use is to find matches between Personen properties to display edges in graph.
    """
    columns_to_check = {
        "address_gmaps": "Adresse",
        "EMailAdresse": "Email",
        "Telefonnummer": "Telefon",
        "VerknuepftesObjektID": "Sonstiges",
    }
    result = pd.DataFrame()
    output_message = []

    for col, match_type in columns_to_check.items():
        if col != "VerknuepftesObjektID":
            # We are merging the dataframe with itself on the column to find internal matches
            temp = df.merge(df, on=col, suffixes=("", "_matched"))
            temp = temp[
                temp["ReferenceID"] < temp["ReferenceID_matched"]
            ]  # To ensure a single match
            temp = temp.dropna(
                subset=[col]
            )  # drop rows where the matching column is NaN

            # New: Additional check for 'address_gmaps'
            # Additional check for 'address_gmaps'
            if col == "address_gmaps":
                mismatched_addresses = temp[
                    ~(
                        (temp["Address1"] == temp["Address1_matched"])
                        | (
                            pd.isna(temp["Address1"])
                            & pd.isna(temp["Address1_matched"])
                        )
                    )
                    | ~(
                        (temp["Address2"] == temp["Address2_matched"])
                        | (
                            pd.isna(temp["Address2"])
                            & pd.isna(temp["Address2_matched"])
                        )
                    )
                ]

                for index, row in mismatched_addresses.iterrows():
                    if not (
                        row["Address1"] == row["Address1_matched"]
                        or (
                            pd.isna(row["Address1"])
                            and pd.isna(row["Address1_matched"])
                        )
                    ):
                        output_message.append(
                            f"Addresszeile 1: \"{row['Address1']}\" in {row['ReferenceID'][-3:]}, aber \"{row['Address1_matched']}\" in {row['ReferenceID_matched'][-3:]}"
                        )

                    if not (
                        row["Address2"] == row["Address2_matched"]
                        or (
                            pd.isna(row["Address2"])
                            and pd.isna(row["Address2_matched"])
                        )
                    ):
                        output_message.append(
                            f"Addresszeile 2: \"{row['Address2']}\" in {row['ReferenceID'][-3:]}, aber \"{row['Address2_matched']}\" in {row['ReferenceID_matched'][-3:]}"
                        )

                temp = temp[
                    (
                        (temp["Address1"] == temp["Address1_matched"])
                        | (
                            pd.isna(temp["Address1"])
                            & pd.isna(temp["Address1_matched"])
                        )
                    )
                    & (
                        (temp["Address2"] == temp["Address2_matched"])
                        | (
                            pd.isna(temp["Address2"])
                            & pd.isna(temp["Address2_matched"])
                        )
                    )
                ]

            temp["match_type"] = match_type
            temp.rename(
                columns={"ReferenceID": "source", "ReferenceID_matched": "ReferenceID"},
                inplace=True,
            )
            result = pd.concat([result, temp], axis=0).drop_duplicates(
                subset=["source", "ReferenceID"]
            )
        else:
            for _, row in df.iterrows():
                lst = convert_string_to_list(row[col])
                for item in lst:
                    mask = (df["ReferenceID"] == item) & (
                        df["ReferenceID"] != row["ReferenceID"]
                    )
                    temp = df[mask].copy()
                    if (
                        not temp.empty
                        and row["ReferenceID"] < temp["ReferenceID"].values[0]
                    ):  # Check if temp is not empty and then ensure a single match
                        temp["match_type"] = match_type
                        temp["source"] = row["ReferenceID"]
                        result = pd.concat([result, temp], axis=0).drop_duplicates()

    return result, output_message


def get_stammdaten_info(combined_organisations, stammdaten):
    unique_nodes = combined_organisations.drop_duplicates(
        subset="ReferenceID", keep="first"
    )
    stammdaten["Mandant"] = stammdaten["Mandant"].replace("eGov UVEK", "UVEK")

    stammdaten = stammdaten.merge(unique_nodes, on="ReferenceID", how="inner")

    stammdaten["AnzahlVerknüpfungen"] = stammdaten["AnzahlVerknüpfungen"].fillna("0")
    stammdaten["AnzahlVerknüpfungen"] = (
        stammdaten["AnzahlVerknüpfungen"].astype(int).astype(str)
    )
    stammdaten["Color"] = stammdaten["Mandant"].apply(
        lambda x: "#63FFFC" if x == "UVEK" else "#9674E9"
    )

    ids = stammdaten["ReferenceID"].tolist()
    labels = list(
        zip(
            stammdaten["Mandant"] + ": " + stammdaten["AnzahlVerknüpfungen"],
            stammdaten["Color"],
        )
    )

    # Ensuring that labels corresponding to one ID are sorted alphabetically, e.g. BAKOM always before UVEK
    sorted_pairs = sorted(zip(ids, labels), key=lambda x: (x[0], x[1]))
    ids, labels = zip(*sorted_pairs)

    return ids, labels, stammdaten


def obtain_uvek_matches(combined_organisations, df_uvek_matches):
    reference_ids = combined_organisations["ReferenceID"]
    # Use the isin() method to filter df_uvek_matches based on the values in reference_ids
    filtered_uvek_matches = df_uvek_matches[
        df_uvek_matches["matches"].isin(reference_ids)
    ]
    filtered_uvek_matches = filtered_uvek_matches.drop_duplicates(subset="ReferenceID")
    return filtered_uvek_matches


def get_true_lists(df):
    """
    Warnung: Speichern zu excel verwandelt Zellen die eine Liste als value enthalten in Strings: ['a','b'] --> "['a','b']"
    Hier fixt dies indem es Verknüpfungsart und VerknüpftesObjektID in Listen umwandelt.
    Muss bei jedem einlesen von xlsx daten neu ausgeführt werden.
    """

    def convert_to_list(s):
        try:
            return ast.literal_eval(s)
        except (ValueError, SyntaxError):
            print(f"error with: {s}")
            return []

    df["Verknuepfungsart_list"] = df["Verknuepfungsart"].apply(convert_to_list)
    df["VerknuepftesObjektID_list"] = df["VerknuepftesObjektID"].apply(convert_to_list)
    return df


def get_true_lists_generic(df):
    """
    Warnung: Speichern zu excel verwandelt Zellen die eine Liste als value enthalten in Strings: ['a','b'] --> "['a','b']"
    Hier fixt dies indem es jede Kolonne die im obigen string format ist umwandelt und in "kolonnename_list" speichert.
    Muss bei jedem einlesen von xlsx daten neu ausgeführt werden.
    """

    def convert_to_list(s):
        try:
            return ast.literal_eval(s)
        except (ValueError, SyntaxError):
            return s  # return the original value if conversion fails

    for col in df.columns:
        # Skip columns that already have a '_list' suffix
        if col.endswith("_list"):
            continue

        # Check if any element in the column is a string that looks like a list
        if (
            df[col]
            .apply(
                lambda x: isinstance(x, str) and x.startswith("[") and x.endswith("]")
            )
            .any()
        ):
            new_col_name = f"{col}_list"
            df[new_col_name] = df[col].apply(convert_to_list)

    return df
