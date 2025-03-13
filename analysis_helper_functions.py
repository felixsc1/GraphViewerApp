import pandas as pd
from hardcoded_values import produkte_dict_name_first
import multiprocessing
from functools import partial
import streamlit as st
import string
import os

def abbreviate_first_name(name):
    name = " ".join(name.split())  # Remove extra spaces
    # name = name.title()  # Standardize case
    parts = name.split()
    if len(parts) > 1 and not parts[0].endswith('.'):
        parts[0] = parts[0][0] + '.'  # Replace the first name with its abbreviation
    return ' '.join(parts)


def find_name_adresse_doubletten(df, organisationen=True, abbreviated_first_name=False, only_with_Geschaeftspartner=False):
    """
    A cluster here is just any group of organizations with exact match in Name and Adresse (email irrelevant). Used for Doubletten analyses.
    """
    # Group by 'Name' and 'address_full', and assign cluster_id
    if organisationen:
        df['cluster_id'] = df.groupby(['Name_Zeile2', 'address_full']).ngroup()
    else:
        if abbreviated_first_name:
            df["Name_abbrev"] = df["Name"].apply(abbreviate_first_name)
            df['cluster_id'] = df.groupby(['Name_abbrev', 'address_full']).ngroup()
        else:
            df['cluster_id'] = df.groupby(['Name', 'address_full']).ngroup()

    # Keep only groups with at least 2 identical rows
    df = df[df.groupby('cluster_id')['cluster_id'].transform('size') > 1]
    
    # If only_with_Geschaeftspartner is True, filter clusters
    if only_with_Geschaeftspartner:
        df = df[df.groupby('cluster_id')['Geschaeftspartner_list'].transform(lambda x: x.str.len().max() > 0)]

    if organisationen:
        df.sort_values(by='Name_Zeile2', inplace=True)
    else:
        df.sort_values(by='Name', inplace=True)
    df.reset_index(drop=True, inplace=True)

    return df


def filter_clusters_with_mixed_produkt_roles(
    df, no_Geschaeftspartner=True, no_Servicerole=True
):
    """
    Für DB cleanup step 2, nachdem alle Produktrollen auf einen Master übertragen wurden.
    Findet Doubletten, bei denen eine Inhaber+Adressant ist (sollte nur noch eine  =Master sein), die anderen Doubletten habe keinerlei Rollen mehr.
    Achtung: Cluster kann weitere Doubletten enthalten, die noch Rollen haben (z.B. nur Rechnungsempfänger), diese werden nicht angezeigt.
    """
    # Apply initial filters to remove entries based on no_Geschaeftspartner and no_Servicerole
    if no_Geschaeftspartner:
        df = df[~df["Geschaeftspartner_list"].apply(lambda gp: len(gp) != 0)]
    if no_Servicerole:
        df = df[df["Servicerole_count"] == 0]

    def filter_relevant_members(group):
        # Identify members with both roles > 0
        has_both_roles = group[
            (group["Produkt_Inhaber"] > 0) & (group["Produkt_Adressant"] > 0)
        ]
        # Identify members with both roles == 0
        has_zero_roles = group[
            (group["Produkt_Inhaber"] == 0) & (group["Produkt_Adressant"] == 0)
        ]
        # Return relevant members if criteria are met
        if len(has_both_roles) == 1 and len(has_zero_roles) >= 1:
            return pd.concat([has_both_roles, has_zero_roles])
        return pd.DataFrame()

    # Apply the group filter and ensure each group has at least two members
    filtered_df = df.groupby("cluster_id").apply(filter_relevant_members).reset_index(drop=True)
    filtered_df = filtered_df[filtered_df.groupby("cluster_id")["cluster_id"].transform("size") >= 2]

    return filtered_df



## CLEANUP / STYLING FUNCTIONS

from openpyxl import load_workbook
from openpyxl.styles import PatternFill
def apply_excel_styling_organisationen_nur_master_hat_produkte(file_path, sheet_name='Sheet1'):
    # Load the workbook and select the sheet
    wb = load_workbook(file_path)
    ws = wb[sheet_name]

    # Define the tomato fill using the RGB part of the color code
    tomato_fill = PatternFill(start_color="FF6347", end_color="FF6347", fill_type="solid")

    # Find the column indices for 'master', 'Produkt_Inhaber', and 'Produkt_Adressant'
    header = {cell.value: idx for idx, cell in enumerate(ws[1])}
    master_col = header.get('master')
    produkt_inhaber_col = header.get('Produkt_Inhaber')
    produkt_adressant_col = header.get('Produkt_Adressant')

    if master_col is None or produkt_inhaber_col is None or produkt_adressant_col is None:
        print("Header values:", [cell.value for cell in ws[1]])
        raise ValueError("One or more required columns are missing in the Excel file.")

    # Iterate over the rows and apply the tomato fill where necessary
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        if (row[master_col].value == 'X' and 
            row[produkt_inhaber_col].value == 0 and 
            row[produkt_adressant_col].value == 0):
            for cell in row:
                cell.fill = tomato_fill

    # Save the workbook
    wb.save(file_path)


def set_master_flag(df):
    if "cluster_id" not in df.columns:
        raise ValueError("DataFrame does not contain 'cluster_id' column.")

    # Create a copy to avoid modifying the original DataFrame
    result_df = df.copy()

    # Sort by 'score' and 'CreatedAt' in descending order for processing
    sorted_df = result_df.sort_values(by=["score", "CreatedAt"], ascending=[False, False])

    # Group by 'cluster_id' and mark the first row in each group as 'X' and others as ''
    sorted_df["master"] = sorted_df.groupby("cluster_id").cumcount().map({0: "X"}).fillna("")

    # Handle the case where all rows have the same 'cluster_id'
    if sorted_df["master"].eq("").all():
        sorted_df["master"].iloc[0] = "X"  # Mark the first row as 'master'

    # Assign the 'master' column directly to the original DataFrame
    result_df["master"] = sorted_df["master"]
    
    # Create the 'master_ID' column
    master_ids = sorted_df[sorted_df["master"] == "X"].set_index("cluster_id")["ReferenceID"]
    result_df["masterID"] = result_df["cluster_id"].map(master_ids)

    return result_df

def renumber_and_sort_alphanumeric(df, column='cluster_id'):    # Split the column into two
    # Split the column into two
    df['num'] = df[column].str.split('_').str[0]
    df['alpha'] = df[column].str.split('_').str[1]
    df['num'] = df['num'].astype(int)

    # Sort the dataframe
    df.sort_values(['num', 'alpha'], inplace=True)

    # Create a new column with rank
    df['rank'] = df['num'].rank(method='dense').astype(int)

    # Combine 'rank' and 'alpha' to get the desired format
    df[column] = df['rank'].astype(str) + '_' + df['alpha']

    # Drop the temporary columns
    df.drop(['num', 'alpha', 'rank'], axis=1, inplace=True)

    return df

def renumber_pairs(column):
    # To recognize clusters I have a continous cluster-id, e.g. 1,1,2,2,3,3, but due to filtering there are some gaps, 1,1,3,3, ...
    # this will just renumber it to 1,1,2,2
    # numbers dont even have to be in ascending order, e.g. if i sorted dataframe by name before and cluster_ids become 3,3,1,1,2,2
    # will just respect the order of appearance and re-number starting with 1.
    mapping = {}
    new_id = 1

    for val in column:
        if val not in mapping:
            mapping[val] = new_id
            new_id += 1

    new_values = [mapping[val] for val in column]
    return new_values

def filter_verknuepfungen(row):
    """
    For Organisations-analyses we only want to show Verknüpfungen to Personen. 
    This filters out anything from the list that is not 'Mitarbeiter' or 'Administrator'
    """
    filtered_indices = [i for i, x in enumerate(row['Verknuepfungsart_list']) if x in ['Mitarbeiter', 'Administrator']]
    row['Verknuepfungsart_list'] = [row['Verknuepfungsart_list'][i] for i in filtered_indices]
    row['VerknuepftesObjektID_list'] = [row['VerknuepftesObjektID_list'][i] for i in filtered_indices]
    row['VerknuepftesObjekt_list'] = [row['VerknuepftesObjekt_list'][i] for i in filtered_indices]
    return row
    
def final_touch(df, cols_to_keep, two_roles=False, alphanumeric=False):
    """
    For a single dataframe. 
    newer analyses have alphanumeric cluster_ids (1_a, 1_b, 2_a, etc.), so we need to sort them differently.
    """
    if "cluster_id" not in df.columns:
        raise ValueError("DataFrame does not contain 'cluster_id' column.")
    
    df = df.apply(filter_verknuepfungen, axis=1)
    df = set_master_flag(df)
    df = df[cols_to_keep]
    df["score"] = df["score"].astype(int)
    
    if two_roles or alphanumeric:
        df = renumber_and_sort_alphanumeric(df.copy(), column='cluster_id')
    else:
        df["cluster_id"] = df["cluster_id"].astype(int)
        df["cluster_id"] = renumber_pairs(df["cluster_id"])
        df.sort_values(by="cluster_id", inplace=True)
        
    df.reset_index(drop=True, inplace=True)
    return df


def final_touch_batch(df_dict, cols_to_keep, two_roles=False, alphanumeric=False):
    """
    Processes any number of dataframes at once.
    Expects a dictionary, with the Description as key and the dataframe or nested dictionary of dataframes as value, as well as the columns to keep.
    Returns a dictionary with key = description and value = dataframe or nested dictionary of dataframes.
    """
    result_dict = {}
    for produktname, value in df_dict.items():
        if isinstance(value, dict):  # Check if the value is a nested dictionary
            nested_result = {}
            for name, df in value.items():
                nested_result[name] = final_touch(df, cols_to_keep, two_roles, alphanumeric=alphanumeric)
            result_dict[produktname] = nested_result
        else:  # If the value is a single dataframe
            result_dict[produktname] = final_touch(value, cols_to_keep, two_roles, alphanumeric=alphanumeric)

    return result_dict


## --- Functions related to finding and filtering Organisationsdoubletten


def general_exclusion_criteria(
    df, no_Produkte=False, no_Geschaeftspartner=True, no_Servicerole=True, only_with_Geschaeftspartner=False
):
    # Apply conditions to each row
    if no_Produkte:
        df = df[((df["Produkt_Inhaber"] == 0) & (df["Produkt_Adressant"] == 0))]
    if no_Geschaeftspartner:
        df = df[~df["Geschaeftspartner_list"].apply(lambda gp: len(gp) != 0)]
    if no_Servicerole:
        df = df[df["Servicerole_count"] == 0]
        
    # Filters on the group level
    # At least one member of a cluster must have a Geschaeftspartner connection
    if only_with_Geschaeftspartner:
        df = df.groupby("cluster_id").filter(
            lambda x: x["Geschaeftspartner_list"].apply(lambda gp: len(gp) > 0).any()
        )

    # Filter out groups with less than 2 members
    return df[df.groupby("cluster_id")["cluster_id"].transform("size") >= 2]


def batch_process_produkte(df, organisationsrollen_df, produktnamen):
    """
    First calls get_product_information() to get Produkt information for every row (slow part).
    Then calls two variations of cleanup functions that check if any group of Doubletten has either all 3 roles for a product or only 2 roles.
    After this we still have a single dataframe for one product. Further processing below is to split it up into different "muster".
    """
    result_3 = {}
    result_2 = {}
    for produktname in produktnamen:
        product_info = get_product_information(df, organisationsrollen_df, produktname)
        st.write(f"✅ Done with {produktname}")
        cleaned_product_info_3 = cleanup_produkte_columns(product_info)
        cleaned_product_info_2 = cleanup_produkte_columns_only_2_roles(product_info)

        if cleaned_product_info_3.empty:
            st.write(f"&nbsp;&nbsp;&nbsp;❌ No Doubletten with 3 roles found!")
        else:
            result_3[produktname] = cleaned_product_info_3

        if cleaned_product_info_2.empty:
            st.write("&nbsp;&nbsp;&nbsp;❌ No Doubletten with 2 roles found!", unsafe_allow_html=True)
        else:
            result_2[produktname] = cleaned_product_info_2
            
    return result_3, result_2

def get_product_information(df_input, organisationsrollen_df, produktname):
    # uses parallel processing to utilize 100% CPU, takes several minutes for one produkt.

    # Pre-filter organisationsrollen_df by produktname
    full_id = produkte_dict_name_first.get(produktname, None)
    if not full_id:
        raise ValueError(f"Produkt '{produktname}' not found in produkte_dict")

    filtered_organisationsrollen_df = organisationsrollen_df[
        organisationsrollen_df["FullID"] == full_id
    ]

    grouped_df = df_input.groupby("cluster_id")
    result_df = parallel_apply(
        grouped_df, worker_function, filtered_organisationsrollen_df
    ).reset_index(drop=True)

    return result_df

def cleanup_produkte_columns(df):
    """
    To be executed after add_singular_produkte_columns_group()
    input df still has groups with empty lists in Inhaber_Objekt etc., which will be removed here.
    Also if any group member has e.g. a produkt listed as Inhaber, but that produkt is nowhere listed as korrempf/rechempf in the group, discards the whole group.
    It does not however care how the products are distributed (e.g. all roles for one org, or distributed across 3)
    """

    def check_group(group):
        # Check if Inhaber_Objekt and Korrempf_Objekt columns are empty in the entire group
        if (
            group["Inhaber_Objekt"].apply(len).sum() == 0
            and group["Korrempf_Objekt"].apply(len).sum() == 0
        ):
            return False  # Mark group for removal

        # Flatten the lists in each column and count occurrences
        all_inhaber = [item for sublist in group["Inhaber_Objekt"] for item in sublist]
        all_korrempf = [
            item for sublist in group["Korrempf_Objekt"] for item in sublist
        ]

        all_rechempf = [
            item for sublist in group["Rechempf_Objekt"] for item in sublist
        ]
        unique_items = set(all_inhaber + all_rechempf + all_korrempf)

        # Check if each item appears in the required columns
        for item in unique_items:
            if all_inhaber.count(item) != 1 or all_korrempf.count(item) != 1:
                return False  # Mark group for removal

            if all_rechempf.count(item) != 1:
                return False  # Mark group for removal if Rechempf_Objekt is considered

        return True  # Keep the group

    # Apply the check to each group and filter the DataFrame
    filtered_df = df.groupby("cluster_id").filter(check_group)

    return filtered_df


def cleanup_produkte_columns_only_2_roles(df):
    """
    Modified to filter out groups where an element appears in exactly two of the three roles.
    Groups where all three roles are empty lists or where an element only appears under one role or in all three roles are removed.
    """

    def check_group(group):
        # Check if all three columns are empty lists in the entire group
        if (
            all(group["Inhaber_Objekt"].apply(len) == 0)
            and all(group["Korrempf_Objekt"].apply(len) == 0)
            and all(group["Rechempf_Objekt"].apply(len) == 0)
        ):
            return False  # Mark group for removal

        # Flatten the lists in each column
        all_inhaber = [item for sublist in group["Inhaber_Objekt"] for item in sublist]
        all_korrempf = [
            item for sublist in group["Korrempf_Objekt"] for item in sublist
        ]
        all_rechempf = [
            item for sublist in group["Rechempf_Objekt"] for item in sublist
        ]

        unique_items = set(all_inhaber + all_rechempf + all_korrempf)

        # Check if each item appears in exactly two of the three roles
        for item in unique_items:
            count = 0
            if item in all_inhaber:
                count += 1
            if item in all_korrempf:
                count += 1
            if item in all_rechempf:
                count += 1

            if count != 2:  # Remove group if item is in 1 or 3 roles
                return False

        return True  # Keep the group if all items are in exactly two roles

    # Apply the check to each group and filter the DataFrame
    filtered_df = df.groupby("cluster_id").filter(check_group)

    return filtered_df


def parallel_apply(
    grouped_df, func, filtered_organisationsrollen_df, num_processes=None
):
    # Create a list of tuples (group_key, group_data)
    group_list = [(group, data) for group, data in grouped_df]

    # Partial function with fixed arguments
    partial_func = partial(
        func, filtered_organisationsrollen_df=filtered_organisationsrollen_df
    )

    # Create a pool of processes
    pool = multiprocessing.Pool(processes=num_processes)

    # Map the function to the chunks and collect the results
    result_chunks = pool.map(partial_func, group_list)

    # Close the pool and wait for work to finish
    pool.close()
    pool.join()

    # Concatenate the results
    return pd.concat(result_chunks)


def worker_function(group_data, filtered_organisationsrollen_df):
    group, data = group_data
    return add_singular_produkte_columns_group_simplified(
        data, filtered_organisationsrollen_df
    )
    
    
def add_singular_produkte_columns_group_simplified(
    group, filtered_organisationsrollen_df
):
    # if organisationsrollen_df is already pre-filtered

    roles = [
        "Inhaber_RefID",
        "Rechnungsempfaenger_RefID",
        "Korrespondenzempfaenger_RefID",
    ]
    role_columns = {
        "Inhaber_RefID": ("Inhaber_Objekt", "Inhaber_ProduktID"),
        "Rechnungsempfaenger_RefID": ("Rechempf_Objekt", "Rechempf_ProduktID"),
        "Korrespondenzempfaenger_RefID": ("Korrempf_Objekt", "Korrempf_ProduktID"),
    }

    # Initialize new columns if they don't exist
    for column in role_columns.values():
        for col in column:
            if col not in group.columns:
                group[col] = [[] for _ in range(len(group))]

    # Iterate over each row in the group
    for index, row in group.iterrows():
        ref_id = row["ReferenceID"]

        # Filter rows in filtered_organisationsrollen_df based on ref_id and roles
        filtered_df = filtered_organisationsrollen_df[
            filtered_organisationsrollen_df[roles].apply(
                lambda x: ref_id in x.values, axis=1
            )
        ]

        for role in roles:
            # Filter for the specific role
            role_filtered_df = filtered_df[filtered_df[role] == ref_id]

            # Update the group DataFrame
            if not role_filtered_df.empty:
                objekt_col, produktid_col = role_columns[role]
                group.at[index, objekt_col] = role_filtered_df["ProduktObj"].tolist()
                group.at[index, produktid_col] = role_filtered_df[
                    "Produkt_RefID"
                ].tolist()

    return group


def organisationsrollen_filter_and_format_batch(
    df_dict, rows_per_product=2, roles_per_product=3
):
    """
    Processes any number of Produkte at once.
    Expects a dictionary, with the Produktname as key and the dataframe as value.
    Returns a dictionary with key = Produktname and value = list of dataframes (Inhaber_separat, etc.)
    """
    result_dict = {}
    statistics_data = []
    
    for key, df in df_dict.items():
        df = set_master_flag(df)  # is needed for re-ordering
        dataframes, names = organisationsrollen_filter_and_format(
            df, rows_per_product=rows_per_product, roles_per_product=roles_per_product
        )

        # Check if dataframes and names lists are not empty and have the same length
        if dataframes and names and len(dataframes) == len(names):
            nested_dict = {
                names[i]: dataframe for i, dataframe in enumerate(dataframes)
            }
            result_dict[key] = nested_dict
        else:
            result_dict[key] = {}  # Or some other placeholder if no data is present
    
    # Calculate statistics
        doubletten_count = sum(df["cluster_id"].nunique() for df in dataframes)
        statistics_data.append({"produkte": key, "Doubletten": doubletten_count})

    # Create statistics DataFrame
    statistics_df = pd.DataFrame(statistics_data)
    return result_dict, statistics_df


def organisationsrollen_filter_and_format(
    input_df, roles_per_product=3, rows_per_product=2
):
    """
    Update to commented-out version above: no longer need special reordering/format step.
    No longer accepts cluster_size parameter.
    main function for organizing Organisationsrollenanalysen. Calls the functions above.
    - Organizes the clusters into several new dataframes (split_produkte_groups())
    """
    if roles_per_product == 3:
        if rows_per_product == 2:
            df_list, df_list_names = split_produkte_groups(input_df)
        elif rows_per_product == 3:
            df_list, df_list_names = split_produkte_groups_komplette_doublette(input_df)
    elif roles_per_product == 2:
        df_list, df_list_names = split_produkte_groups_two_roles(input_df)
    else:
        raise ValueError("roles_per_product must be 2 or 3")

    return df_list, df_list_names


def split_produkte_groups(df):
    """
    Commented-out version above works, but was meant for clusters of size 2 and has outdated numbering and filtering scheme.
    Checks if a a group contains all roles for the given product.
    Discards group members that have no role (or that have all three roles).
    Organizes them into list with 3 dataframes.
    """
    df_inhaber_separat = pd.DataFrame(columns=df.columns)
    df_rechempf_separat = pd.DataFrame(columns=df.columns)
    df_korrempf_separat = pd.DataFrame(columns=df.columns)

    alphabet = list(string.ascii_lowercase)
    suffixes = alphabet + [i + j for i in alphabet for j in alphabet]

    for cluster_id, group in df.groupby("cluster_id"):
        unique_elements = set()
        for member in group.itertuples():
            for col in ["Inhaber_Objekt", "Rechempf_Objekt", "Korrempf_Objekt"]:
                unique_elements.update(getattr(member, col))

        sorted_elements = sorted(map(str, unique_elements))
        suffix_dict = {
            elem: f"_{suffixes[i]}" for i, elem in enumerate(sorted_elements)
        }

        for element in unique_elements:
            element_roles = {
                col: set()
                for col in ["Inhaber_Objekt", "Rechempf_Objekt", "Korrempf_Objekt"]
            }
            for member in group.itertuples():
                for col in element_roles.keys():
                    if element in getattr(member, col):
                        element_roles[col].add(member.Index)

            # Ensure roles are distributed across exactly two members and all three roles are covered
            if (
                sum(len(roles) for roles in element_roles.values()) == 3
                and len({idx for roles in element_roles.values() for idx in roles}) == 2
            ):
                # Determine which member has which role(s)
                member_roles_count = {
                    member: sum(member in roles for roles in element_roles.values())
                    for member in group.index
                }

                # Find the member with a single role and two roles
                single_role_member = next(
                    (
                        member
                        for member, count in member_roles_count.items()
                        if count == 1
                    ),
                    None,
                )
                two_roles_member = next(
                    (
                        member
                        for member, count in member_roles_count.items()
                        if count == 2
                    ),
                    None,
                )

                # This checks which dataframe to append based on the role of the single_role_member
                if single_role_member is not None and two_roles_member is not None:
                    for role, members in element_roles.items():
                        if single_role_member in members:
                            single_role = role
                            break

                    for member in group.itertuples():
                        if member.Index in [single_role_member, two_roles_member]:
                            new_row = {
                                col: getattr(member, col)
                                for col in df.columns
                                if col
                                not in [
                                    "Inhaber_Objekt",
                                    "Rechempf_Objekt",
                                    "Korrempf_Objekt",
                                    "Inhaber_ProduktID",
                                    "Rechempf_ProduktID",
                                    "Korrempf_ProduktID",
                                ]
                            }
                            new_row["cluster_id"] = (
                                f"{cluster_id}{suffix_dict[str(element)]}"
                            )

                            for col, id_col in zip(
                                [
                                    "Inhaber_Objekt",
                                    "Rechempf_Objekt",
                                    "Korrempf_Objekt",
                                ],
                                [
                                    "Inhaber_ProduktID",
                                    "Rechempf_ProduktID",
                                    "Korrempf_ProduktID",
                                ],
                            ):
                                if element in getattr(member, col):
                                    index = getattr(member, col).index(element)
                                    new_row[col] = element
                                    new_row[id_col] = (
                                        getattr(member, id_col)[index]
                                        if index < len(getattr(member, id_col))
                                        else ""
                                    )
                                else:
                                    new_row[col] = ""
                                    new_row[id_col] = ""

                            new_row_df = pd.DataFrame([new_row])

                            # Assign to the correct dataframe based on the single_role
                            if single_role == "Inhaber_Objekt":
                                df_inhaber_separat = pd.concat(
                                    [df_inhaber_separat, new_row_df], ignore_index=True
                                )
                            elif single_role == "Rechempf_Objekt":
                                df_rechempf_separat = pd.concat(
                                    [df_rechempf_separat, new_row_df], ignore_index=True
                                )
                            elif single_role == "Korrempf_Objekt":
                                df_korrempf_separat = pd.concat(
                                    [df_korrempf_separat, new_row_df], ignore_index=True
                                )

    df_list = []
    df_list_names = []
    if not df_inhaber_separat.empty:
        df_list.append(df_inhaber_separat)
        df_list_names.append("Inhaber_Separat")
    if not df_rechempf_separat.empty:
        df_list.append(df_rechempf_separat)
        df_list_names.append("Rechempf_Separat")
    if not df_korrempf_separat.empty:
        df_list.append(df_korrempf_separat)
        df_list_names.append("KorrEmpf_Separat")

    return df_list, df_list_names


def split_produkte_groups_komplette_doublette(df):
    """
    Checks if cluster_id groups of input df have all three roles for a product for one member each.
    Discards additional members that have no role or groups that don't match the criteria.
    """
    # List columns that are expected to contain lists and their corresponding ProduktID columns
    objekt_columns = ["Inhaber_Objekt", "Rechempf_Objekt", "Korrempf_Objekt"]
    produktid_columns = [
        "Inhaber_ProduktID",
        "Rechempf_ProduktID",
        "Korrempf_ProduktID",
    ]
    list_columns = objekt_columns + produktid_columns

    # Initialize the two DataFrames
    df_komplette_doubletten = pd.DataFrame(columns=df.columns)
    df_sonstige = pd.DataFrame(columns=df.columns)

    alphabet = list(string.ascii_lowercase)
    suffixes = alphabet + [i + j for i in alphabet for j in alphabet]

    for cluster_id, group in df.groupby("cluster_id"):
        unique_elements = set()
        for member in group.itertuples():
            for col in objekt_columns:
                unique_elements.update(getattr(member, col))

        sorted_elements = sorted(map(str, unique_elements))
        suffix_dict = {
            elem: f"_{suffixes[i]}" for i, elem in enumerate(sorted_elements)
        }

        for element in unique_elements:
            element_roles = {col: set() for col in objekt_columns}
            for member in group.itertuples():
                for col in objekt_columns:
                    if element in getattr(member, col):
                        element_roles[col].add(member.Index)

            for member in group.itertuples():
                roles_per_member = {
                    member.Index: len(
                        [
                            col
                            for col in objekt_columns
                            if member.Index in element_roles[col]
                        ]
                    )
                    for member in group.itertuples()
                }
                if (
                    len(
                        [
                            role_count
                            for role_count in roles_per_member.values()
                            if role_count == 1
                        ]
                    )
                    == 3
                ):
                    # Check if member has a role for the given element
                    if any(
                        member.Index in element_roles[col] for col in objekt_columns
                    ):
                        new_row = {
                            col: getattr(member, col)
                            for col in df.columns
                            if col not in list_columns
                        }
                        new_row["cluster_id"] = (
                            f"{cluster_id}{suffix_dict[str(element)]}"
                        )

                        for col, id_col in zip(objekt_columns, produktid_columns):
                            if element in getattr(member, col):
                                index = getattr(member, col).index(element)
                                new_row[col] = element
                                new_row[id_col] = (
                                    getattr(member, id_col)[index]
                                    if index < len(getattr(member, id_col))
                                    else ""
                                )
                            else:
                                new_row[col] = ""
                                new_row[id_col] = ""

                        new_row_df = pd.DataFrame([new_row])
                        df_komplette_doubletten = pd.concat(
                            [df_komplette_doubletten, new_row_df], ignore_index=True
                        )

    df_list = []
    df_list_names = []
    if not df_komplette_doubletten.empty:
        df_list.append(df_komplette_doubletten)
        df_list_names.append("Komplette_Doubletten")

    return df_list, df_list_names


def split_produkte_groups_two_roles(df):
    """
    Organizes the DataFrame into four new DataFrames based on common elements in two of the three roles.
    Each unique element within a cluster_id group will produce two rows with the same suffix.
    Members without a role are removed (although they are still doubletten, but must be handled elsewhere).
    Corresponding ProduktID for each Objekt element is also placed in the new row.
    """
    # List columns that are expected to contain lists and their corresponding ProduktID columns
    objekt_columns = ["Inhaber_Objekt", "Rechempf_Objekt", "Korrempf_Objekt"]
    produktid_columns = [
        "Inhaber_ProduktID",
        "Rechempf_ProduktID",
        "Korrempf_ProduktID",
    ]
    list_columns = objekt_columns + produktid_columns

    # Initialize the four DataFrames
    df_inhaber_korrempf = pd.DataFrame(columns=df.columns)
    df_inhaber_rechempf = pd.DataFrame(columns=df.columns)
    df_korrempf_rechempf = pd.DataFrame(columns=df.columns)
    df_sonstige = pd.DataFrame(columns=df.columns)

    alphabet = list(string.ascii_lowercase)
    suffixes = alphabet + [i + j for i in alphabet for j in alphabet]

    for cluster_id, group in df.groupby("cluster_id"):
        unique_elements = set()
        for member in group.itertuples():
            for col in objekt_columns:
                unique_elements.update(getattr(member, col))

        # suffix_dict = {elem: f"_{suffixes[i]}" for i, elem in enumerate(sorted(unique_elements))}
        sorted_elements = sorted(map(str, unique_elements))
        suffix_dict = {
            elem: f"_{suffixes[i]}" for i, elem in enumerate(sorted_elements)
        }

        for element in unique_elements:
            element_roles = {col: set() for col in objekt_columns}
            for member in group.itertuples():
                for col in objekt_columns:
                    if element in getattr(member, col):
                        element_roles[col].add(member.Index)

            for member in group.itertuples():
                new_row = {
                    col: getattr(member, col)
                    for col in df.columns
                    if col not in list_columns
                }
                new_row["cluster_id"] = f"{cluster_id}{suffix_dict[str(element)]}"

                for col, id_col in zip(objekt_columns, produktid_columns):
                    if element in getattr(member, col):
                        index = getattr(member, col).index(element)
                        new_row[col] = element
                        new_row[id_col] = (
                            getattr(member, id_col)[index]
                            if index < len(getattr(member, id_col))
                            else ""
                        )
                    else:
                        new_row[col] = ""
                        new_row[id_col] = ""

                new_row_df = pd.DataFrame([new_row])

                # only process rows that actually have a role:
                if any(new_row[col] for col in objekt_columns):
                    # Determine the DataFrame to which the row should be assigned
                    if (
                        len(element_roles["Inhaber_Objekt"]) > 1
                        and len(element_roles["Rechempf_Objekt"]) > 1
                    ):
                        df_sonstige = pd.concat(
                            [df_sonstige, new_row_df], ignore_index=True
                        )
                    elif (
                        len(element_roles["Inhaber_Objekt"]) > 1
                        and len(element_roles["Korrempf_Objekt"]) > 1
                    ):
                        df_sonstige = pd.concat(
                            [df_sonstige, new_row_df], ignore_index=True
                        )
                    elif (
                        len(element_roles["Rechempf_Objekt"]) > 1
                        and len(element_roles["Korrempf_Objekt"]) > 1
                    ):
                        df_sonstige = pd.concat(
                            [df_sonstige, new_row_df], ignore_index=True
                        )
                    elif (
                        len(element_roles["Inhaber_Objekt"]) > 0
                        and len(element_roles["Rechempf_Objekt"]) > 0
                    ):
                        if new_row["Inhaber_Objekt"] and new_row["Rechempf_Objekt"]:
                            df_sonstige = pd.concat(
                                [df_sonstige, new_row_df], ignore_index=True
                            )
                        else:
                            df_inhaber_rechempf = pd.concat(
                                [df_inhaber_rechempf, new_row_df], ignore_index=True
                            )
                    elif (
                        len(element_roles["Inhaber_Objekt"]) > 0
                        and len(element_roles["Korrempf_Objekt"]) > 0
                    ):
                        if new_row["Inhaber_Objekt"] and new_row["Korrempf_Objekt"]:
                            df_sonstige = pd.concat(
                                [df_sonstige, new_row_df], ignore_index=True
                            )
                        else:
                            df_inhaber_korrempf = pd.concat(
                                [df_inhaber_korrempf, new_row_df], ignore_index=True
                            )
                    elif (
                        len(element_roles["Rechempf_Objekt"]) > 0
                        and len(element_roles["Korrempf_Objekt"]) > 0
                    ):
                        if new_row["Rechempf_Objekt"] and new_row["Korrempf_Objekt"]:
                            df_sonstige = pd.concat(
                                [df_sonstige, new_row_df], ignore_index=True
                            )
                        else:
                            df_korrempf_rechempf = pd.concat(
                                [df_korrempf_rechempf, new_row_df], ignore_index=True
                            )
                    else:
                        df_sonstige = pd.concat(
                            [df_sonstige, new_row_df], ignore_index=True
                        )

    # Ensure there are no single-member groups in df_sonstige
    df_sonstige = df_sonstige[df_sonstige.duplicated(subset="cluster_id", keep=False)]
    df_korrempf_rechempf = df_korrempf_rechempf[
        df_korrempf_rechempf.duplicated(subset="cluster_id", keep=False)
    ]
    df_inhaber_rechempf = df_inhaber_rechempf[
        df_inhaber_rechempf.duplicated(subset="cluster_id", keep=False)
    ]
    df_inhaber_korrempf = df_inhaber_korrempf[
        df_inhaber_korrempf.duplicated(subset="cluster_id", keep=False)
    ]

    df_list = []
    df_list_names = []
    if not df_inhaber_korrempf.empty:
        df_list.append(df_inhaber_korrempf)
        df_list_names.append("Inhaber_KorrEmpf")
    if not df_inhaber_rechempf.empty:
        df_list.append(df_inhaber_rechempf)
        df_list_names.append("Inhaber_RechEmpf")
    if not df_korrempf_rechempf.empty:
        df_list.append(df_korrempf_rechempf)
        df_list_names.append("KorrEmpf_RechEmpf")
    if not df_sonstige.empty:
        df_list.append(df_sonstige)
        df_list_names.append("Sonstige")

    return df_list, df_list_names


def create_excel_files_from_nested_dict(nested_dict, output_dir="output"):
    """
    For output of Organisationsrollenanalyse:
    - top level keys are suffixes of the .xlsx files
    - keys of nested dicts are sheet names

    Error message "ValueError: seek of closed file" is expected and can be ignored
    """
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    for file_suffix, sheet_dict in nested_dict.items():
        if any(not df.empty for df in sheet_dict.values()):
            file_name = f"{output_dir}/Organisationen_{file_suffix}.xlsx"
            with pd.ExcelWriter(file_name, engine="openpyxl") as writer:
                for sheet_name, df in sheet_dict.items():
                    if not df.empty:
                        df.to_excel(writer, sheet_name=sheet_name, index=False)
                    else:
                        print(
                            f"Empty dataframe for file '{file_suffix}', sheet '{sheet_name}'"
                        )
        else:
            print(
                f"No sheets created for file '{file_suffix}' as all dataframes are empty."
            )


if __name__ == "__main__":
    """
    Doesnt really matter what is placed here, prevents some error message if main block were missing.
    """
    def get_product_information(df_input, organisationsrollen_df, produktname):
        # uses parallel processing to utilize 100% CPU, takes several minutes for one produkt.

        # Pre-filter organisationsrollen_df by produktname
        full_id = produkte_dict_name_first.get(produktname, None)
        if not full_id:
            raise ValueError(f"Produkt '{produktname}' not found in produkte_dict")

        filtered_organisationsrollen_df = organisationsrollen_df[
            organisationsrollen_df["FullID"] == full_id
        ]

        grouped_df = df_input.groupby("cluster_id")
        result_df = parallel_apply(
            grouped_df, worker_function, filtered_organisationsrollen_df
        ).reset_index(drop=True)

        return result_df